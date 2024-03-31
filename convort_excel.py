import openpyxl
from PIL import Image as pilimg
from openpyxl import Workbook
from openpyxl.drawing.image import Image

def convert_info_to_excel(file_path, listPicture, listInfo,idx):
    # Create a new Workbook object
    try:
        workbook = openpyxl.load_workbook('example.xlsx')
    except FileNotFoundError:
        workbook = Workbook()
    # Select the active worksheet
    ws = workbook.active

    # Fill some cells with data
    ws['A1'] = 'picture'
    ws['B1'] = "Supplier's reference"
    ws['C1'] = "Supplier's designation "
    ws['D1'] = 'PRODUCT RANGE'
    ws['E1'] = 'Colour(s)'
    ws['F1'] = 'Measure units'
    ws['G1'] = 'Brand/License'
    ws['H1'] = 'BIUB or BBD*\n(dd/mm/yyyy)'
    ws['I1'] = 'Untaxed (Wine)'
    ws['J1'] = 'Qty available'
    ws['K1'] = 'Wholesale Price'
    ws['L1'] = 'Clearance Price'
    ws['M1'] = 'Retail price'
    ws['N1'] = 'Packing details'
    ws['O1'] = 'Nb packets / pallet'
    ws['P1'] = 'Number of pallets'
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 40
    ws.column_dimensions['G'].width = 40
    ws.column_dimensions['H'].width = 40
    ws.column_dimensions['I'].width = 40
    ws.column_dimensions['J'].width = 40
    ws.column_dimensions['K'].width = 40
    ws.column_dimensions['L'].width = 40
    ws.column_dimensions['M'].width = 40
    ws.column_dimensions['N'].width = 40
    ws.column_dimensions['O'].width = 40
    ws.column_dimensions['P'].width = 40
    for i in range(len(listInfo)):
        print(listInfo[i])

        k=i+2+idx
        img2=pilimg.open(listPicture[i])
        img2=img2.resize((100,100))
        img2.save(listPicture[i])
        img = Image(listPicture[i])
        #img = img.resize((100, 100))  # Resize the image to 100x100
        # Add the image to the worksheet
        ws.add_image(img, "A"+ str(k))
        ws.row_dimensions[k].height = 80

        ws["B" + str(k)] = listInfo[i][0].split("-")[0] + "-" + listInfo[i][0].split("-")[1]

        ws["C" + str(k)] = listInfo[i][1]


        ws["D" + str(k)] = "ACCESSORIES"
        ws["E" + str(k)] = (listInfo[i][1].split())[-1]

        ws["F" + str(k)] = "One size fits most"

        ws["G" + str(k)] = "C.C."
        ws["J" + str(k)] = listInfo[i][2].split(":")[1]


    # Save the workbook
    workbook.save(file_path)
'''
file_path = 'example.xlsx'

list_info = [['DA-201-BLK', 'KNIT PATCH HAT WITH POM - BLACK', 'Qty On-Hand: 27'],
             ['DA-201-GRY', 'KNIT PATCH HAT WITH POM - GREY', 'Qty On-Hand: 17'],
             ['DA-201-NVY', 'KNIT PATCH HAT WITH POM - NAVY', 'Qty On-Hand: 33'],
             ['DA-201-RED', 'KNIT PATCH HAT WITH POM - RED', 'Qty On-Hand: 28'],
             ['DA-201-WHT', 'KNIT PATCH HAT WITH POM - WHITE', 'Qty On-Hand: 27'],
             ['G-303-BLACK', 'KNIT MITTEN WITH REAL FUR CUFF', 'Qty On-Hand: 120']]

convert_info_to_excel(file_path, [], list_info)
'''